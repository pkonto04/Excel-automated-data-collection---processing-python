"""
consolidate_app.py
------------------
Double-click to run. No terminal needed.
To build as .exe:
    pip install pyinstaller
    pyinstaller --onefile --windowed consolidate_app.py
"""

import os
import threading
from collections import defaultdict
from tkinter import *
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar, Style

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

# ── Questions — edit these to match your real questionnaire ──────────────────
QUESTIONS = [
    {"id": "Q1",  "text": "Question 1",  "type": "dropdown", "options": ["Option A", "Option B", "Option C"]},
    {"id": "Q2",  "text": "Question 2",  "type": "dropdown", "options": ["Yes", "No", "N/A"]},
    {"id": "Q3",  "text": "Question 3",  "type": "number"},
    {"id": "Q4",  "text": "Question 4",  "type": "date"},
    {"id": "Q5",  "text": "Question 5",  "type": "dropdown", "options": ["Low", "Medium", "High"]},
    {"id": "Q6",  "text": "Question 6",  "type": "text"},
    {"id": "Q7",  "text": "Question 7",  "type": "dropdown", "options": ["Option A", "Option B", "Option C", "Option D"]},
    {"id": "Q8",  "text": "Question 8",  "type": "number"},
    {"id": "Q9",  "text": "Question 9",  "type": "dropdown", "options": ["Yes", "No"]},
    {"id": "Q10", "text": "Question 10", "type": "text"},
]

DROPDOWN_QS = [q for q in QUESTIONS if q["type"] == "dropdown"]

# ── Excel styles ─────────────────────────────────────────────────────────────
HDR_FONT  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HDR_FILL  = PatternFill("solid", start_color="1F4E79")
ALT_FILL  = PatternFill("solid", start_color="EBF3FB")
ERR_FILL  = PatternFill("solid", start_color="FCE4D6")
OK_FILL   = PatternFill("solid", start_color="E2EFDA")
NORM_FONT = Font(name="Arial", size=11)
BOLD_FONT = Font(name="Arial", bold=True, size=11)
XL_CENTER = Alignment(horizontal="center", vertical="center")
XL_LEFT   = Alignment(horizontal="left",   vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

def style_header(cell):
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = XL_CENTER
    cell.border = THIN_BORDER

def style_cell(cell, alt=False):
    cell.font = NORM_FONT
    cell.fill = ALT_FILL if alt else PatternFill("solid", start_color="FFFFFF")
    cell.alignment = XL_LEFT
    cell.border = THIN_BORDER

# ── Core logic ────────────────────────────────────────────────────────────────
def read_submission(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    def get_named(name):
        if name in wb.defined_names:
            dest = list(wb.defined_names[name].destinations)
            if dest:
                sheet_name, coord = dest[0]
                return wb[sheet_name][coord].value
        return None

    client = get_named("client_name") or ws["C3"].value
    answers = {}
    for i, q in enumerate(QUESTIONS):
        val = get_named(f"answer_{q['id']}")
        if val is None:
            val = ws.cell(row=6 + i, column=3).value
        answers[q["id"]] = val

    wb.close()
    return {"client": client or os.path.basename(filepath), "answers": answers}

def write_responses(wb, records):
    ws = wb.create_sheet("All Responses")
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 28
    for i in range(len(QUESTIONS)):
        col_letter = chr(66 + i) if i < 26 else "A" + chr(65 + i - 26)
        ws.column_dimensions[col_letter].width = 22

    headers = ["Client"] + [f"{q['id']}: {q['text']}" for q in QUESTIONS]
    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=col, value=h))
    ws.row_dimensions[1].height = 22

    for r, record in enumerate(records, start=2):
        alt = r % 2 == 0
        c = ws.cell(row=r, column=1, value=record["client"])
        c.font = BOLD_FONT
        c.fill = ALT_FILL if alt else PatternFill("solid", start_color="FFFFFF")
        c.alignment = XL_LEFT
        c.border = THIN_BORDER
        for col, q in enumerate(QUESTIONS, start=2):
            val = record["answers"].get(q["id"])
            cell = ws.cell(row=r, column=col, value=val)
            style_cell(cell, alt)
            if q["type"] == "date" and val:
                cell.number_format = "DD/MM/YYYY"

def write_validation(wb, records):
    ws = wb.create_sheet("Validation")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 40
    for col, h in enumerate(["Client", "Question", "Issue", "Detail"], start=1):
        style_header(ws.cell(row=1, column=col, value=h))

    row = 2
    dropdown_valid = {q["id"]: q["options"] for q in QUESTIONS if q["type"] == "dropdown"}
    total_issues = 0

    for record in records:
        for q in QUESTIONS:
            val = record["answers"].get(q["id"])
            issue = detail = None
            if val is None or str(val).strip() == "":
                issue, detail = "Missing answer", "Cell was left blank"
            elif q["type"] == "number":
                try:
                    float(val)
                except (TypeError, ValueError):
                    issue, detail = "Invalid type", f'Expected a number, got "{val}"'
            elif q["type"] == "dropdown" and str(val) not in dropdown_valid[q["id"]]:
                issue = "Invalid option"
                detail = f'"{val}" not in: {", ".join(dropdown_valid[q["id"]])}'
            if issue:
                total_issues += 1
                ws.cell(row=row, column=1, value=record["client"]).font = NORM_FONT
                ws.cell(row=row, column=2, value=q["id"]).font = NORM_FONT
                c = ws.cell(row=row, column=3, value=issue)
                c.font = Font(name="Arial", size=11, color="C00000")
                c.fill = ERR_FILL
                ws.cell(row=row, column=4, value=detail).font = NORM_FONT
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = THIN_BORDER
                row += 1

    if total_issues == 0:
        c = ws.cell(row=2, column=1, value="No issues found — all submissions are valid.")
        c.font = Font(name="Arial", bold=True, size=11, color="375623")
        c.fill = OK_FILL

def write_statistics(wb, records):
    ws = wb.create_sheet("Statistics")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    current_row = 1

    for q in DROPDOWN_QS:
        counts = defaultdict(int)
        total = 0
        for record in records:
            val = record["answers"].get(q["id"])
            if val and str(val).strip():
                counts[str(val).strip()] += 1
                total += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        hdr = ws.cell(row=current_row, column=1, value=f"{q['id']}: {q['text']}")
        hdr.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        hdr.fill = HDR_FILL
        hdr.alignment = XL_LEFT
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for col, label in enumerate(["Option", "Count", "% of responses"], start=1):
            c = ws.cell(row=current_row, column=col, value=label)
            c.font = Font(name="Arial", bold=True, size=11)
            c.fill = PatternFill("solid", start_color="D6E4F0")
            c.border = THIN_BORDER
            c.alignment = XL_CENTER
        current_row += 1

        data_start = current_row
        for option in q["options"]:
            count = counts.get(option, 0)
            pct = f"{round(count / total * 100, 1)}%" if total > 0 else "0%"
            for col, val in enumerate([option, count, pct], start=1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.font = NORM_FONT
                cell.border = THIN_BORDER
                cell.alignment = XL_CENTER
            current_row += 1

        data_end = current_row - 1
        chart = BarChart()
        chart.type = "col"
        chart.title = f"{q['id']}: {q['text']}"
        chart.y_axis.title = "Count"
        chart.style = 10
        chart.width = 14
        chart.height = 9
        chart.legend = None
        chart.add_data(Reference(ws, min_col=2, min_row=data_start, max_row=data_end))
        chart.set_categories(Reference(ws, min_col=1, min_row=data_start, max_row=data_end))
        ws.add_chart(chart, f"E{data_start - 1}")
        current_row += 2

def run_consolidation(folder, log, progress, run_btn):
    files = [
        os.path.join(folder, f)
        for f in sorted(os.listdir(folder))
        if f.endswith(".xlsx") and not f.startswith("~")
    ]

    if not files:
        log("No .xlsx files found in the selected folder.")
        run_btn.config(state=NORMAL)
        return

    records = []
    progress["maximum"] = len(files) + 1

    for i, f in enumerate(files):
        try:
            record = read_submission(f)
            records.append(record)
            log(f"✓  {os.path.basename(f)}  —  {record['client']}")
        except Exception as e:
            log(f"✗  {os.path.basename(f)}  —  ERROR: {e}")
        progress["value"] = i + 1

    home = os.path.expanduser("~")
    onedrive_desktop = os.path.join(home, "OneDrive", "Desktop")
    standard_desktop = os.path.join(home, "Desktop")
    desktop = onedrive_desktop if os.path.exists(onedrive_desktop) else standard_desktop
    os.makedirs(desktop, exist_ok=True)
    output_path = os.path.join(desktop, "consolidated_results.xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    write_responses(wb, records)
    write_validation(wb, records)
    write_statistics(wb, records)
    wb.save(output_path)
    progress["value"] = len(files) + 1

    log("")
    log(f"  Done!  Processed {len(records)} file(s).")
    log(f"  Saved to: {output_path}")
    run_btn.config(state=NORMAL)

# ── GUI ───────────────────────────────────────────────────────────────────────
class App:
    def __init__(self, root):
        root.title("Questionnaire Consolidator")
        root.resizable(False, False)
        root.configure(bg="#F5F5F5")

        DARK   = "#1F4E79"
        MID    = "#2E75B6"
        LIGHT  = "#F5F5F5"
        WHITE  = "#FFFFFF"
        FONT   = ("Arial", 10)
        FONT_B = ("Arial", 10, "bold")

        # ── Header ────────────────────────────────────────────────────────────
        header = Frame(root, bg=DARK, pady=16)
        header.pack(fill=X)
        Label(header, text="Questionnaire Consolidator",
              font=("Arial", 14, "bold"), bg=DARK, fg=WHITE).pack()
        Label(header, text="Merge all bank submissions into one Excel file",
              font=("Arial", 9), bg=DARK, fg="#BDD7EE").pack()

        # ── Body ──────────────────────────────────────────────────────────────
        body = Frame(root, bg=LIGHT, padx=24, pady=20)
        body.pack(fill=BOTH)

        Label(body, text="Submissions folder", font=FONT_B,
              bg=LIGHT, fg="#2F2F2F").grid(row=0, column=0, sticky=W, pady=(0, 4))

        folder_frame = Frame(body, bg=LIGHT)
        folder_frame.grid(row=1, column=0, sticky=EW, pady=(0, 16))

        self.folder_var = StringVar(value="No folder selected")
        folder_entry = Entry(folder_frame, textvariable=self.folder_var,
                             font=FONT, width=46, relief=FLAT,
                             bg=WHITE, fg="#595959",
                             highlightthickness=1, highlightbackground="#D9D9D9")
        folder_entry.pack(side=LEFT, ipady=6, padx=(0, 8))

        browse_btn = Button(folder_frame, text="Browse…", font=FONT_B,
                            bg=MID, fg=WHITE, relief=FLAT,
                            activebackground=DARK, activeforeground=WHITE,
                            cursor="hand2", padx=12, pady=6,
                            command=self.browse)
        browse_btn.pack(side=LEFT)

        # Progress bar
        style = Style()
        style.theme_use("default")
        style.configure("blue.Horizontal.TProgressbar",
                        troughcolor="#E0E0E0", background=MID,
                        thickness=8)
        self.progress = Progressbar(body, style="blue.Horizontal.TProgressbar",
                                    orient=HORIZONTAL, length=500, mode="determinate")
        self.progress.grid(row=2, column=0, sticky=EW, pady=(0, 12))

        # Log box
        log_frame = Frame(body, bg=WHITE, highlightthickness=1,
                          highlightbackground="#D9D9D9")
        log_frame.grid(row=3, column=0, sticky=EW, pady=(0, 16))

        self.log_text = Text(log_frame, font=("Courier New", 9),
                             height=10, width=62, relief=FLAT,
                             bg=WHITE, fg="#2F2F2F",
                             state=DISABLED, wrap=WORD)
        scrollbar = Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.pack(side=LEFT, fill=BOTH, expand=True, padx=8, pady=8)
        scrollbar.pack(side=RIGHT, fill=Y)

        # Run button
        self.run_btn = Button(body, text="▶  Run Consolidation",
                              font=("Arial", 11, "bold"),
                              bg=DARK, fg=WHITE, relief=FLAT,
                              activebackground=MID, activeforeground=WHITE,
                              cursor="hand2", padx=20, pady=10,
                              command=self.run)
        self.run_btn.grid(row=4, column=0)

        # Footer
        footer = Frame(root, bg="#E8E8E8", pady=8)
        footer.pack(fill=X)
        Label(footer, text="Output is saved automatically to your Desktop",
              font=("Arial", 8), bg="#E8E8E8", fg="#888888").pack()

        root.update_idletasks()
        w, h = root.winfo_width(), root.winfo_height()
        x = (root.winfo_screenwidth()  - w) // 2
        y = (root.winfo_screenheight() - h) // 2
        root.geometry(f"+{x}+{y}")

    def browse(self):
        folder = filedialog.askdirectory(title="Select folder containing filled questionnaires")
        if folder:
            self.folder_var.set(folder)
            self.log(f"Folder selected: {folder}")

    def log(self, msg):
        self.log_text.config(state=NORMAL)
        self.log_text.insert(END, msg + "\n")
        self.log_text.see(END)
        self.log_text.config(state=DISABLED)

    def run(self):
        folder = self.folder_var.get()
        if folder == "No folder selected" or not os.path.isdir(folder):
            messagebox.showwarning("No folder", "Please select a valid submissions folder first.")
            return

        self.run_btn.config(state=DISABLED)
        self.progress["value"] = 0
        self.log_text.config(state=NORMAL)
        self.log_text.delete("1.0", END)
        self.log_text.config(state=DISABLED)
        self.log(f"Starting consolidation...\n")

        thread = threading.Thread(
            target=run_consolidation,
            args=(folder, self.log, self.progress, self.run_btn),
            daemon=True
        )
        thread.start()

if __name__ == "__main__":
    root = Tk()
    app = App(root)
    root.mainloop()
